from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from .data_ingest import (
    load_metrics_json_file,
    normalize_metric_hints,
    structured_metrics_to_document_contexts,
)
from .documents import (
    COMPANY_HINTS,
    _extract_metric_hints,
    detect_companies_from_text,
    extract_metric_hints_for_company,
    parse_pdf_document,
)

COMPANY_COLUMNS = frozenset(
    {"company", "company_name", "ticker", "symbol", "name", "公司", "企业"}
)
METRIC_NAME_COLUMNS = frozenset({"metric", "metrics", "indicator", "field", "指标", "科目"})
METRIC_VALUE_COLUMNS = frozenset({"value", "amount", "数值", "值"})

SUPPORTED_SUFFIXES = frozenset(
    {".pdf", ".csv", ".xlsx", ".md", ".markdown", ".json", ".txt"}
)


def _normalize_header(name: str) -> str:
    return re.sub(r"\s+", "_", (name or "").strip().lower())


def _build_document_context(
    *,
    document_id: str,
    filename: str,
    text: str,
    source_type: str,
    path: str | None = None,
    pages: list[str] | None = None,
    detected_companies: list[str] | None = None,
    metric_hints: dict[str, float] | None = None,
    tables: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    page_list = pages if pages is not None else [text]
    companies = detected_companies or detect_companies_from_text(text, filename)
    hints = metric_hints if metric_hints is not None else {}
    per_company = {
        company: extract_metric_hints_for_company(text, company) for company in companies
    }
    ctx: dict[str, Any] = {
        "document_id": document_id,
        "filename": filename,
        "text": text,
        "pages": page_list,
        "page_count": len(page_list),
        "excerpt": text[:4000],
        "detected_companies": companies,
        "metric_hints": hints,
        "per_company_metric_hints": per_company,
        "source_type": source_type,
    }
    if path is not None:
        ctx["path"] = path
    if tables:
        ctx["tables"] = tables
    return ctx


def _metrics_from_row(row: dict[str, Any]) -> tuple[str | None, dict[str, Any]]:
    company: str | None = None
    metrics: dict[str, Any] = {}
    for key, value in row.items():
        if key is None or value is None or str(value).strip() == "":
            continue
        norm = _normalize_header(str(key))
        if norm in COMPANY_COLUMNS:
            company = str(value).strip()
            continue
        metrics[str(key).strip()] = value
    return company, metrics


def _pivot_long_format_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Convert metric/value rows into wide rows grouped by company."""
    if not rows:
        return rows
    headers = {_normalize_header(str(key)) for row in rows for key in row.keys()}
    if not (headers & METRIC_NAME_COLUMNS and headers & METRIC_VALUE_COLUMNS):
        return rows

    metric_key = next(
        (key for row in rows for key in row if _normalize_header(str(key)) in METRIC_NAME_COLUMNS),
        None,
    )
    value_key = next(
        (key for row in rows for key in row if _normalize_header(str(key)) in METRIC_VALUE_COLUMNS),
        None,
    )
    if metric_key is None or value_key is None:
        return rows

    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        company = None
        for key, value in row.items():
            if _normalize_header(str(key)) in COMPANY_COLUMNS:
                company = str(value).strip()
                break
        metric_name = row.get(metric_key)
        metric_value = row.get(value_key)
        if metric_name is None or metric_value is None or str(metric_name).strip() == "":
            continue
        bucket_key = company or "__default__"
        grouped.setdefault(bucket_key, {})[str(metric_name).strip()] = metric_value

    wide_rows: list[dict[str, Any]] = []
    for company, metrics in grouped.items():
        row = dict(metrics)
        if company != "__default__":
            row["company"] = company
        wide_rows.append(row)
    return wide_rows


def _context_from_company_metrics(
    company: str,
    metrics: dict[str, Any],
    *,
    filename: str,
    source_type: str,
    path: str | None = None,
    document_id_suffix: str = "",
) -> dict[str, Any]:
    hints = normalize_metric_hints(metrics)
    serialized = json.dumps(metrics, ensure_ascii=False)
    doc_id = f"structured_{company}{document_id_suffix}".replace(" ", "_")
    return _build_document_context(
        document_id=doc_id,
        filename=filename,
        text=serialized,
        source_type=source_type,
        path=path,
        pages=[serialized],
        detected_companies=[company],
        metric_hints=hints,
        tables=[{"company": company, "metrics": metrics}],
    )


def _parse_tabular_rows(
    rows: list[dict[str, Any]],
    *,
    file_path: Path,
    source_type: str,
    default_company: str | None = None,
) -> list[dict[str, Any]]:
    if not rows:
        raise ValueError(f"No data rows found in {file_path.name}")

    rows = _pivot_long_format_rows(rows)

    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        company, metrics = _metrics_from_row(row)
        if not metrics:
            continue
        resolved = company or default_company or file_path.stem
        bucket = grouped.setdefault(resolved, {})
        bucket.update(metrics)

    if not grouped:
        raise ValueError(f"No recognizable metric columns in {file_path.name}")

    contexts: list[dict[str, Any]] = []
    for company, metrics in grouped.items():
        contexts.append(
            _context_from_company_metrics(
                company,
                metrics,
                filename=file_path.name,
                source_type=source_type,
                path=str(file_path),
            )
        )
    return contexts


def _read_csv_rows(file_path: Path) -> list[dict[str, Any]]:
    raw = file_path.read_text(encoding="utf-8-sig")
    if not raw.strip():
        raise ValueError(f"CSV file is empty: {file_path.name}")
    reader = csv.DictReader(raw.splitlines())
    if reader.fieldnames is None:
        raise ValueError(f"CSV file has no header row: {file_path.name}")
    return [dict(row) for row in reader]


def parse_csv_documents(file_path: Path) -> list[dict[str, Any]]:
    rows = _read_csv_rows(file_path)
    return _parse_tabular_rows(rows, file_path=file_path, source_type="csv")


def parse_excel_documents(file_path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "Excel upload requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    contexts: list[dict[str, Any]] = []
    try:
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            values = list(sheet.iter_rows(values_only=True))
            if not values:
                continue
            headers = [str(cell).strip() if cell is not None else "" for cell in values[0]]
            if not any(headers):
                continue
            rows: list[dict[str, Any]] = []
            for raw_row in values[1:]:
                if raw_row is None or not any(cell is not None and str(cell).strip() for cell in raw_row):
                    continue
                row = {
                    headers[idx]: raw_row[idx]
                    for idx in range(min(len(headers), len(raw_row)))
                    if headers[idx]
                }
                rows.append(row)
            if not rows:
                continue
            default_company = sheet_name.strip() if sheet_name.strip().lower() != "sheet1" else file_path.stem
            contexts.extend(
                _parse_tabular_rows(
                    rows,
                    file_path=file_path,
                    source_type="excel",
                    default_company=default_company,
                )
            )
    finally:
        workbook.close()

    if not contexts:
        raise ValueError(f"No usable sheets or metrics found in {file_path.name}")
    return contexts


def _split_markdown_sections(text: str) -> list[str]:
    sections = [
        section.strip()
        for section in re.split(r"(?=^##\s+)", text, flags=re.MULTILINE)
        if section.strip()
    ]
    return sections or [text]


def parse_markdown_document(file_path: Path) -> dict[str, Any]:
    text = file_path.read_text(encoding="utf-8").strip()
    if not text:
        raise ValueError(f"Markdown file is empty: {file_path.name}")
    pages = _split_markdown_sections(text)
    return _build_document_context(
        document_id=file_path.stem,
        filename=file_path.name,
        text=text,
        source_type="markdown",
        path=str(file_path),
        pages=pages,
        detected_companies=detect_companies_from_text(text, file_path.name),
        metric_hints=_extract_metric_hints(text),
    )


def parse_text_document(file_path: Path) -> dict[str, Any]:
    text = file_path.read_text(encoding="utf-8").strip()
    if not text:
        raise ValueError(f"Text file is empty: {file_path.name}")
    return _build_document_context(
        document_id=file_path.stem,
        filename=file_path.name,
        text=text,
        source_type="text",
        path=str(file_path),
        pages=[text],
        detected_companies=detect_companies_from_text(text, file_path.name),
        metric_hints=_extract_metric_hints(text),
    )


def parse_json_documents(file_path: Path) -> list[dict[str, Any]]:
    company_metrics = load_metrics_json_file(file_path)
    contexts = structured_metrics_to_document_contexts(company_metrics)
    for ctx in contexts:
        ctx["path"] = str(file_path)
    return contexts


def parse_upload_documents(file_path: Path) -> list[dict[str, Any]]:
    """Route an uploaded file to the correct parser; returns one or more document_context dicts."""
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        supported = ", ".join(sorted(SUPPORTED_SUFFIXES))
        raise ValueError(
            f"Unsupported upload type '{suffix}' for {path.name}. Supported: {supported}"
        )

    if suffix == ".pdf":
        ctx = parse_pdf_document(path)
        ctx["source_type"] = "pdf"
        return [ctx]
    if suffix == ".csv":
        return parse_csv_documents(path)
    if suffix == ".xlsx":
        return parse_excel_documents(path)
    if suffix in {".md", ".markdown"}:
        return [parse_markdown_document(path)]
    if suffix == ".txt":
        return [parse_text_document(path)]
    if suffix == ".json":
        return parse_json_documents(path)
    raise ValueError(f"Unsupported upload type: {suffix}")
